from __future__ import annotations

import os
import hashlib
import json
import csv
import re
import sqlite3
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .consistency_checks import (
    append_consistency_sheet,
    execute_consistency_checks,
    is_consistency_checks_enabled,
)
from .logging_setup import debug_extra
from .models import FileDescriptor, MergedRow, ParsedRow


class SpodPipeline:
    """Главный orchestrator загрузки CSV, консолидации и выгрузки."""

    def __init__(self, config: dict[str, Any], logger: Any) -> None:
        self.config = config
        self.logger = logger
        self.run_id: str = datetime.now().strftime("%Y%m%d_%H%M%S")

        self.paths: dict[str, Path] = {
            "input_root": Path(config["paths"]["input_root"]),
            "output_excel_dir": self._ensure_dated_output_dir(
                Path(config["paths"]["output_excel_dir"])
            ),
            "output_db_dir": Path(config["paths"]["output_db_dir"]),
        }
        self.entities: dict[str, dict[str, Any]] = config["entities"]
        self.stands: list[str] = config["stands"]
        self.entity_field_orders: dict[str, dict[str, list[str]]] = defaultdict(dict)
        self.entity_extra_fields: dict[str, set[str]] = defaultdict(set)
        self._diff_report_rows: list[dict[str, str]] = []
        self.runtime_cfg: dict[str, Any] = config.get("runtime", {})
        self.dry_run: bool = bool(self.runtime_cfg.get("dry_run", False))
        self.parallel_workers: int = self._resolve_parallel_workers()

        self.paths["output_excel_dir"].mkdir(parents=True, exist_ok=True)
        self.paths["output_db_dir"].mkdir(parents=True, exist_ok=True)
        self._consistency_violations: list[Any] = []

        db_name: str = config["sqlite"]["db_name"]
        self.db_path: Path = self.paths["output_db_dir"] / db_name

    def _ensure_dated_output_dir(self, base_dir: Path) -> Path:
        """Возвращает директорию формата YYYY/MM-DD, избегая двойной вложенности."""
        if len(base_dir.parts) >= 2:
            year_part = base_dir.parts[-2]
            month_day_part = base_dir.parts[-1]
            if re.fullmatch(r"\d{4}", year_part) and re.fullmatch(r"\d{2}-\d{2}", month_day_part):
                return base_dir

        now = datetime.now()
        year_dir = now.strftime("%Y")
        month_day_dir = now.strftime("%m-%d")
        return base_dir / year_dir / month_day_dir

    def _resolve_parallel_workers(self) -> int:
        """Определяет число потоков: авто по ядрам или из runtime.parallel_workers."""
        requested = self.runtime_cfg.get("parallel_workers", "auto")
        if isinstance(requested, int) and requested > 0:
            return requested
        cpu_count = os.cpu_count() or 1
        return max(1, min(32, cpu_count))

    def run(self) -> tuple[Path, Path]:
        """Запускает полный цикл обработки по ТЗ."""
        stage_timings: dict[str, float] = {}
        run_started = time.perf_counter()
        self._diff_report_rows = []
        conn = sqlite3.connect(self.db_path)
        try:
            self.logger.info("Проверка входных файлов перед стартом...")
            self.logger.info(
                "Режим выполнения: dry_run=%s, parallel_workers=%s",
                self.dry_run,
                self.parallel_workers,
            )
            self._init_db(conn)
            stage_start = time.perf_counter()
            files = self._scan_files()
            stage_timings["scan_files"] = time.perf_counter() - stage_start
            self._log_found_files_summary(files)
            stage_start = time.perf_counter()
            parsed, parse_stats = self._parse_all_rows(files, conn)
            stage_timings["parse_rows"] = time.perf_counter() - stage_start
            self._log_parse_summary(parse_stats)
            stage_start = time.perf_counter()
            merged, merge_stats, diff_rows = self._merge_rows(parsed)
            stage_timings["merge_rows"] = time.perf_counter() - stage_start
            self._diff_report_rows = diff_rows
            self._log_merge_summary(merge_stats)
            stage_start = time.perf_counter()
            self._run_consistency_checks(parsed, merged)
            stage_timings["consistency_checks"] = time.perf_counter() - stage_start
            stage_start = time.perf_counter()
            self._save_merged_rows(conn, merged)
            stage_timings["save_merged_rows"] = time.perf_counter() - stage_start
            stage_start = time.perf_counter()
            excel_path = self._export_excel(merged)
            stage_timings["export_excel"] = time.perf_counter() - stage_start
            self._save_run(conn, "SUCCESS", excel_path)
            conn.commit()
            self._log_stage_timings(stage_timings, run_started)
            self.logger.info("Обработка завершена успешно")
            return self.db_path, excel_path
        except Exception:
            self._save_run(conn, "FAILED", None)
            conn.commit()
            self._log_stage_timings(stage_timings, run_started)
            raise
        finally:
            conn.close()

    def _init_db(self, conn: sqlite3.Connection) -> None:
        """Создает таблицы SQLite и индексы дедупликации."""
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS runs (
                run_id TEXT PRIMARY KEY,
                started_at TEXT NOT NULL,
                finished_at TEXT,
                result_status TEXT NOT NULL,
                excel_path TEXT,
                log_info_path TEXT,
                log_debug_path TEXT
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS ingested_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id TEXT NOT NULL,
                stand TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                file_path TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_hash_sha256 TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                loaded_at TEXT NOT NULL,
                status TEXT NOT NULL
            )
            """
        )
        cursor.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS ux_ingested_files_hash ON ingested_files(file_hash_sha256)"
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS raw_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id TEXT NOT NULL,
                stand TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                row_num INTEGER NOT NULL,
                row_json TEXT NOT NULL,
                row_hash TEXT NOT NULL
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS merged_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                business_key TEXT NOT NULL,
                row_hash TEXT NOT NULL,
                source_stands TEXT NOT NULL,
                source_count INTEGER NOT NULL,
                is_equal_all INTEGER NOT NULL,
                diff_group_key TEXT NOT NULL,
                merged_json TEXT NOT NULL
            )
            """
        )
        cursor.execute(
            """
            INSERT OR REPLACE INTO runs(run_id, started_at, result_status)
            VALUES(?, ?, ?)
            """,
            (self.run_id, datetime.now().isoformat(), "RUNNING"),
        )

    def _save_run(
        self, conn: sqlite3.Connection, status: str, excel_path: Path | None
    ) -> None:
        cursor = conn.cursor()
        cursor.execute(
            """
            UPDATE runs
            SET finished_at = ?, result_status = ?, excel_path = ?
            WHERE run_id = ?
            """,
            (
                datetime.now().isoformat(),
                status,
                str(excel_path) if excel_path else None,
                self.run_id,
            ),
        )

    def _save_merged_rows(
        self, conn: sqlite3.Connection, merged: dict[str, list[MergedRow]]
    ) -> None:
        """Сохраняет merged-набор пакетно в SQLite (кроме dry-run)."""
        if self.dry_run:
            self.logger.info("Dry-run: запись merged_rows в БД пропущена")
            return
        payload: list[tuple[Any, ...]] = []
        for entity, rows in merged.items():
            for row in rows:
                payload.append(
                    (
                        self.run_id,
                        entity,
                        row.business_key,
                        row.row_hash,
                        row.source_stands,
                        row.source_count,
                        1 if row.is_equal_all else 0,
                        row.diff_group_key,
                        json.dumps(row.merged_data, ensure_ascii=False),
                    )
                )
        if not payload:
            return
        conn.cursor().executemany(
            """
            INSERT INTO merged_rows(
                run_id, entity_type, business_key, row_hash, source_stands,
                source_count, is_equal_all, diff_group_key, merged_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            payload,
        )

    def _scan_files(self) -> list[FileDescriptor]:
        """Сканирует входные каталоги по явным именам файлов из config.json."""
        descriptors: list[FileDescriptor] = []
        input_root: Path = self.paths["input_root"]
        missing_items: list[str] = []
        existing_candidates: list[tuple[str, str, Path]] = []

        for stand in self.stands:
            stand_dir = input_root / stand
            if not stand_dir.exists():
                missing_items.append(f"Каталог стенда отсутствует: {stand_dir}")
                continue

            for entity, details in self.entities.items():
                file_name = details["file_names"][stand]
                file_path = stand_dir / file_name
                if not file_path.exists():
                    missing_items.append(
                        f"Не найден файл entity={entity}, stand={stand}: {file_path}"
                    )
                    continue
                existing_candidates.append((stand, entity, file_path))

        if missing_items:
            message = "Старт прерван. Не найдены обязательные входные данные:\n" + "\n".join(
                f"- {item}" for item in missing_items
            )
            raise FileNotFoundError(message)

        with ThreadPoolExecutor(max_workers=self.parallel_workers) as executor:
            futures = [
                executor.submit(self._build_file_descriptor, stand, entity, file_path)
                for stand, entity, file_path in existing_candidates
            ]
            for future in as_completed(futures):
                descriptors.append(future.result())

        descriptors.sort(key=lambda item: (item.entity, item.stand))

        self.logger.info("Найдено входных файлов: %s", len(descriptors))
        return descriptors

    def _build_file_descriptor(self, stand: str, entity: str, file_path: Path) -> FileDescriptor:
        """Собирает метаданные файла (включая SHA-256)."""
        file_hash = hashlib.sha256(file_path.read_bytes()).hexdigest()
        return FileDescriptor(
            stand=stand,
            entity=entity,
            file_path=file_path,
            file_name=file_path.name,
            file_hash=file_hash,
            file_size=file_path.stat().st_size,
        )

    def _parse_all_rows(
        self, files: list[FileDescriptor], conn: sqlite3.Connection
    ) -> tuple[dict[str, list[ParsedRow]], dict[str, Any]]:
        """Парсит CSV, формирует ключи/хэши строк и пишет данные в SQLite."""
        by_entity: dict[str, list[ParsedRow]] = defaultdict(list)
        cursor = conn.cursor()
        stats: dict[str, Any] = {
            "total_files": len(files),
            "db_new_files": 0,
            "db_skipped_files": 0,
            "processed_rows": 0,
            "rows_by_entity": defaultdict(int),
        }

        file_is_new_hash: dict[str, bool] = {}
        for descriptor in files:
            is_new_hash = self._is_new_file_hash(cursor, descriptor.file_hash)
            file_is_new_hash[descriptor.file_hash] = is_new_hash
            status = "NEW" if is_new_hash else "SKIPPED_DUPLICATE"
            if is_new_hash:
                stats["db_new_files"] += 1
            else:
                stats["db_skipped_files"] += 1
            cursor.execute(
                """
                INSERT OR IGNORE INTO ingested_files(
                    run_id, stand, entity_type, file_path, file_name,
                    file_hash_sha256, file_size, loaded_at, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    self.run_id,
                    descriptor.stand,
                    descriptor.entity,
                    str(descriptor.file_path),
                    descriptor.file_name,
                    descriptor.file_hash,
                    descriptor.file_size,
                    datetime.now().isoformat(),
                    status,
                ),
            )

        with ThreadPoolExecutor(max_workers=self.parallel_workers) as executor:
            futures = [executor.submit(self._parse_file_descriptor, descriptor) for descriptor in files]
            for future in as_completed(futures):
                descriptor, headers, parsed_rows = future.result()
                self.entity_field_orders[descriptor.entity][descriptor.stand] = list(headers)
                by_entity[descriptor.entity].extend(parsed_rows)
                stats["processed_rows"] += len(parsed_rows)
                stats["rows_by_entity"][descriptor.entity] += len(parsed_rows)

                if not self.dry_run:
                    raw_rows_payload = [
                        (
                            self.run_id,
                            descriptor.stand,
                            descriptor.entity,
                            row.row_num,
                            json.dumps(row.data, ensure_ascii=False),
                            row.row_hash,
                        )
                        for row in parsed_rows
                    ]
                    cursor.executemany(
                        """
                        INSERT INTO raw_rows(run_id, stand, entity_type, row_num, row_json, row_hash)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        raw_rows_payload,
                    )

                self.logger.debug(
                    "Обработан файл %s (%s), строк=%s",
                    descriptor.file_name,
                    "NEW" if file_is_new_hash.get(descriptor.file_hash, False) else "SKIPPED_DUPLICATE",
                    len(parsed_rows),
                    extra=debug_extra("SpodPipeline", "_parse_all_rows"),
                )

        return by_entity, stats

    def _parse_file_descriptor(
        self, descriptor: FileDescriptor
    ) -> tuple[FileDescriptor, list[str], list[ParsedRow]]:
        """Парсит один файл и подготавливает ParsedRow-пакет."""
        headers, rows = self._read_csv_rows_preserve_raw(descriptor.file_path)
        parsed_rows: list[ParsedRow] = []
        for row_num, row in enumerate(rows, start=2):
            normalized = self._normalize_row(row)
            business_key = self._build_business_key(descriptor.entity, row)
            row_hash = self._hash_json(normalized)
            parsed_rows.append(
                ParsedRow(
                    stand=descriptor.stand,
                    entity=descriptor.entity,
                    row_num=row_num,
                    data=row,
                    business_key=business_key,
                    row_hash=row_hash,
                )
            )
        # Дубликат business_key в одном файле даёт несколько ParsedRow на стенд; source_stands
        # тогда отражает все стенды, где есть строка с теми же значениями, что в отчёте (любая из дублей).
        key_lines: dict[str, list[int]] = defaultdict(list)
        for pr in parsed_rows:
            key_lines[pr.business_key].append(pr.row_num)
        for bk, lines in key_lines.items():
            if len(lines) > 1:
                self.logger.warning(
                    "В файле %s (%s) business_key повторяется %s раз на строках %s%s",
                    descriptor.file_name,
                    descriptor.stand,
                    len(lines),
                    lines[:15],
                    "..." if len(lines) > 15 else "",
                    extra=debug_extra("SpodPipeline", "_parse_file_descriptor"),
                )
        return descriptor, headers, parsed_rows

    def _read_csv_rows_preserve_raw(
        self, file_path: Path
    ) -> tuple[list[str], list[dict[str, str]]]:
        """Читает CSV без изменения текстового содержимого полей."""
        content = file_path.read_text(encoding="utf-8-sig")
        records: list[str] = []
        buffer: list[str] = []
        in_quotes = False
        i = 0
        while i < len(content):
            ch = content[i]
            if ch == '"':
                buffer.append(ch)
                if i + 1 < len(content) and content[i + 1] == '"':
                    buffer.append(content[i + 1])
                    i += 2
                    continue
                in_quotes = not in_quotes
                i += 1
                continue
            if ch == "\n" and not in_quotes:
                record = "".join(buffer).rstrip("\r")
                if record:
                    records.append(record)
                buffer = []
                i += 1
                continue
            buffer.append(ch)
            i += 1
        tail = "".join(buffer).rstrip("\r")
        if tail:
            records.append(tail)

        if not records:
            return [], []

        raw_headers = self._split_csv_record_preserve_raw(records[0])
        # Единые имена колонок между стендами: BOM/пробелы в заголовке ломают сопоставление
        # с эталонной строкой (ложное совпадение source_stands при разном смысле столбцов).
        headers = [self._normalize_csv_header_name(h) for h in raw_headers]
        rows: list[dict[str, str]] = []
        for record in records[1:]:
            values = self._split_csv_record_preserve_raw(record)
            if len(values) < len(headers):
                values.extend([""] * (len(headers) - len(values)))
            if len(values) > len(headers):
                values = values[: len(headers)]
            rows.append(dict(zip(headers, values)))
        return headers, rows

    def _normalize_csv_header_name(self, name: str) -> str:
        """Убирает BOM и лишние пробелы в имени колонки CSV для согласованности между файлами."""
        cleaned = name.replace("\ufeff", "").strip()
        return cleaned.lstrip("\ufeff").strip()

    def _split_csv_record_preserve_raw(self, record: str) -> list[str]:
        """Разделяет CSV-запись по `;`, сохраняя поле в исходном виде."""
        fields: list[str] = []
        buffer: list[str] = []
        in_quotes = False
        i = 0
        while i < len(record):
            ch = record[i]
            if ch == '"':
                buffer.append(ch)
                if i + 1 < len(record) and record[i + 1] == '"':
                    buffer.append(record[i + 1])
                    i += 2
                    continue
                in_quotes = not in_quotes
                i += 1
                continue
            if ch == ";" and not in_quotes:
                fields.append("".join(buffer))
                buffer = []
                i += 1
                continue
            buffer.append(ch)
            i += 1
        fields.append("".join(buffer))
        return fields

    def _is_new_file_hash(self, cursor: sqlite3.Cursor, file_hash: str) -> bool:
        cursor.execute(
            "SELECT 1 FROM ingested_files WHERE file_hash_sha256 = ? LIMIT 1", (file_hash,)
        )
        return cursor.fetchone() is None

    def _normalize_row(self, row: dict[str, str | None]) -> dict[str, str]:
        """Канонизирует CSV-строку для корректного сравнения между стендами."""
        normalized: dict[str, str] = {}
        trim_values = self.config.get("merge", {}).get("trim_values", False)
        empty_to_null = self.config.get("merge", {}).get("empty_to_null", False)
        for key, value in row.items():
            if value is None:
                normalized[key] = "" if not empty_to_null else "NULL"
                continue
            # По умолчанию сохраняем значения в исходном виде.
            if trim_values:
                normalized[key] = value.strip()
            else:
                normalized[key] = value
            if empty_to_null and normalized[key] == "":
                normalized[key] = "NULL"
        return normalized

    def _build_business_key(self, entity: str, row: dict[str, str]) -> str:
        """Формирует бизнес-ключ по конфигу; fallback — полный hash строки."""
        key_fields: list[str] = self.entities[entity]["business_key"]
        values: list[str] = [row.get(field, "") for field in key_fields]
        if all(value == "" for value in values):
            return f"HASH:{self._hash_json(row)}"
        return "|".join(values)

    def _hash_json(self, payload: dict[str, Any]) -> str:
        canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True)
        return hashlib.sha256(canonical.encode("utf-8")).hexdigest()

    def _merge_rows(
        self, parsed_rows: dict[str, list[ParsedRow]]
    ) -> tuple[dict[str, list[MergedRow]], dict[str, Any], list[dict[str, str]]]:
        """Объединяет данные по business_key + row_hash и формирует служебные признаки."""
        result: dict[str, list[MergedRow]] = {}
        diff_report_rows: list[dict[str, str]] = []
        stats: dict[str, Any] = {
            "merged_total": 0,
            "equal_all_total": 0,
            "different_total": 0,
            "same_key_different_values_total": 0,
            "by_entity": {},
        }

        with ThreadPoolExecutor(max_workers=self.parallel_workers) as executor:
            futures = {
                executor.submit(self._merge_entity_bundle, entity, rows): entity
                for entity, rows in parsed_rows.items()
            }
            for future in as_completed(futures):
                entity, merged_rows, by_business_key_counts, extra_fields, entity_diff_rows = future.result()
                self.entity_extra_fields[entity] = extra_fields
                result[entity] = merged_rows
                diff_report_rows.extend(entity_diff_rows)
                entity_equal_all = sum(1 for item in merged_rows if item.is_equal_all)
                entity_diff = len(merged_rows) - entity_equal_all
                same_key_different = sum(
                    1 for count in by_business_key_counts.values() if count > 1
                )
                stats["by_entity"][entity] = {
                    "merged_rows": len(merged_rows),
                    "equal_all_rows": entity_equal_all,
                    "different_rows": entity_diff,
                    "same_key_different_values": same_key_different,
                }
                stats["merged_total"] += len(merged_rows)
                stats["equal_all_total"] += entity_equal_all
                stats["different_total"] += entity_diff
                stats["same_key_different_values_total"] += same_key_different
        return result, stats, diff_report_rows

    def _run_consistency_checks(
        self,
        parsed: dict[str, list[ParsedRow]],
        merged: dict[str, list[MergedRow]],
    ) -> None:
        """Запускает проверки консистентности из config.json; дополняет merged_data и список нарушений."""
        self._consistency_violations = []
        cc = self.config.get("consistency_checks") or {}
        if not is_consistency_checks_enabled(cc):
            return
        field_orders: dict[str, dict[str, list[str]]] = {
            ent: dict(st_map) for ent, st_map in self.entity_field_orders.items()
        }
        res = execute_consistency_checks(
            config=self.config,
            stands=list(self.stands),
            entities=self.entities,
            parsed_by_entity=parsed,
            merged_by_entity=merged,
            field_orders=field_orders,
            logger=self.logger,
        )
        self._consistency_violations = res.violations
        self.logger.info(
            "Проверки консистентности завершены: отклонений=%s",
            len(self._consistency_violations),
            extra={"class_name": "SpodPipeline", "func_name": "_run_consistency_checks"},
        )
        self._log_consistency_by_rule_and_stand()

    def _log_consistency_by_rule_and_stand(self) -> None:
        """Пишет в лог таблицу: нарушения по правилам в разрезе стендов."""
        if not self._consistency_violations:
            self.logger.info(
                "CONSISTENCY: нарушений нет (таблица по правилам/стендам пустая).",
                extra={"class_name": "SpodPipeline", "func_name": "_log_consistency_by_rule_and_stand"},
            )
            return

        stands = list(self.stands)
        # Нарушения без конкретного стенда (например merged) учитываем отдельно.
        extra_col = "NO_STAND"
        counts: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        rule_types: dict[str, str] = {}

        for v in self._consistency_violations:
            rid = str(v.rule_id)
            rule_types[rid] = str(v.rule_type)
            st = str(v.stand) if v.stand else extra_col
            counts[rid][st] += 1

        columns = stands + [extra_col, "TOTAL"]
        header = ["RULE_ID", "TYPE"] + columns
        rows: list[list[str]] = []
        for rid in sorted(counts.keys()):
            by_stand = counts[rid]
            total = sum(by_stand.values())
            row = [rid, rule_types.get(rid, "")]
            for c in columns:
                if c == "TOTAL":
                    row.append(str(total))
                else:
                    row.append(str(by_stand.get(c, 0)))
            rows.append(row)

        widths = [len(h) for h in header]
        for r in rows:
            for i, cell in enumerate(r):
                widths[i] = max(widths[i], len(cell))

        def _fmt_row(cells: list[str]) -> str:
            return " | ".join(cells[i].ljust(widths[i]) for i in range(len(cells)))

        sep = "-+-".join("-" * w for w in widths)
        self.logger.info(
            "CONSISTENCY: нарушения по правилам и стендам\n%s\n%s\n%s",
            _fmt_row(header),
            sep,
            "\n".join(_fmt_row(r) for r in rows),
            extra={"class_name": "SpodPipeline", "func_name": "_log_consistency_by_rule_and_stand"},
        )

    def _merge_entity_bundle(
        self, entity: str, rows: list[ParsedRow]
    ) -> tuple[str, list[MergedRow], dict[str, int], set[str], list[dict[str, str]]]:
        """Мержит одну сущность с приоритетом стенда: PROM -> PSI -> IFT."""
        optional_cfg = self.entities[entity].get("optional_fields", {})
        extra_fields = set()
        if optional_cfg.get("enabled", False):
            field_orders = self.entity_field_orders.get(entity, {})
            reference_stand = optional_cfg.get("reference_stand", "PROM")
            reference_fields = set(field_orders.get(reference_stand, []))
            all_fields: set[str] = set()
            for stand_fields in field_orders.values():
                all_fields.update(stand_fields)
            extra_fields = all_fields - reference_fields

        merged_rows, by_business_key_counts, diff_rows = self._merge_entity_default(entity, rows)

        for item in merged_rows:
            has_diff_values = bool(str(item.merged_data.get("same_key_diff_stands", "")).strip())
            item.merged_data["same_key_diff_values_flag"] = "1" if has_diff_values else "0"
            item.merged_data["same_key_diff_values_note"] = "YES" if has_diff_values else "NO"
        merged_rows.sort(key=lambda item: (item.business_key, item.row_hash))
        return entity, merged_rows, by_business_key_counts, extra_fields, diff_rows

    def _source_row_matches_display_payload(
        self, display: dict[str, str], source: dict[str, str]
    ) -> bool:
        """
        Проверяет, что в source те же значения по всем колонкам, что выводятся из display (эталон строки).
        Сравнение ячеек — с усечением пробелов по краям (в CSV часто лишние пробелы).
        """
        for key, value in display.items():
            if source.get(key, "").strip() != str(value).strip():
                return False
        return True

    def _collect_stands_matching_display_payload(
        self,
        business_key: str,
        display_payload: dict[str, str],
        candidates: list[ParsedRow],
    ) -> list[str]:
        """
        Стенды, где для того же business_key в CSV есть строка с тем же набором значений полей,
        что и отображаемая merged-строка (не «стенд встречается в партиции», а «на этом стенде
        реально лежат эти значения»).
        """
        rank = self._stand_sort_rank()
        found: set[str] = set()
        for pr in candidates:
            if pr.business_key != business_key:
                continue
            if self._source_row_matches_display_payload(display_payload, pr.data):
                found.add(pr.stand)
        return sorted(found, key=lambda stand: rank.get(stand, 99))

    def _merge_entity_default(
        self, entity: str, rows: list[ParsedRow]
    ) -> tuple[list[MergedRow], dict[str, int], list[dict[str, str]]]:
        by_key_rows: dict[str, list[ParsedRow]] = defaultdict(list)
        for row in rows:
            by_key_rows[row.business_key].append(row)
        return self._merge_by_business_key_clusters(entity, by_key_rows)

    def _raw_row_fingerprint(self, data: dict[str, str]) -> str:
        """Стабильный отпечаток сырой строки для разделения групп с одинаковым row_hash."""
        canonical = json.dumps(dict(sorted(data.items())), ensure_ascii=False, separators=(",", ":"))
        return hashlib.sha256(canonical.encode("utf-8")).hexdigest()

    def _stand_sort_rank(self) -> dict[str, int]:
        """Порядок стендов для сортировки source_stands (эталонный стенд — первый)."""
        merge_cfg = self.config.get("merge", {})
        preferred = str(merge_cfg.get("reference_row_stand", "PROM"))
        order: list[str] = []
        if preferred in self.stands:
            order.append(preferred)
        for stand in self.stands:
            if stand not in order:
                order.append(stand)
        return {stand: idx for idx, stand in enumerate(order)}

    def _pick_reference_parsed_row(self, rows: list[ParsedRow]) -> ParsedRow:
        """Выбирает строку-эталон для вывода значений в merged-строке."""
        merge_cfg = self.config.get("merge", {})
        preferred = str(merge_cfg.get("reference_row_stand", "PROM"))
        rank = self._stand_sort_rank()
        preferred_rows = [row for row in rows if row.stand == preferred]
        if preferred_rows:
            return preferred_rows[0]
        return min(rows, key=lambda row: rank.get(row.stand, 999))

    def _merge_by_business_key_clusters(
        self, entity: str, by_key_rows: dict[str, list[ParsedRow]]
    ) -> tuple[list[MergedRow], dict[str, int], list[dict[str, str]]]:
        """Выбирает одну строку на ключ по приоритету стенда и собирает детализацию в DIFF_REPORT."""
        merged_rows: list[MergedRow] = []
        by_business_key_counts: dict[str, int] = defaultdict(int)
        diff_report_rows: list[dict[str, str]] = []
        rank = self._stand_sort_rank_for_duplicates()

        for business_key, key_rows in by_key_rows.items():
            by_business_key_counts[business_key] = len(key_rows)
            selected_row = self._select_priority_row_for_key(key_rows)
            same_stands, diff_stands = self._split_stands_by_selected_row(selected_row, key_rows)
            merged_payload = dict(selected_row.data)
            effective_row_hash = self._hash_json(self._normalize_row(merged_payload))
            diff_group_key = f"{entity}:{business_key}:1"

            source_stands = "-".join(same_stands)
            merged_payload["source_stands"] = source_stands
            merged_payload["source_count"] = str(len(same_stands))
            merged_payload["is_equal_all"] = "1" if len(same_stands) == len(self.stands) else "0"
            merged_payload["diff_group_key"] = diff_group_key
            merged_payload["same_row_stands"] = "-".join(same_stands)
            merged_payload["same_key_diff_stands"] = "-".join(diff_stands)
            merged_rows.append(
                MergedRow(
                    entity=entity,
                    business_key=business_key,
                    row_hash=effective_row_hash,
                    source_stands=source_stands,
                    source_count=len(same_stands),
                    is_equal_all=len(same_stands) == len(self.stands),
                    diff_group_key=diff_group_key,
                    merged_data=merged_payload,
                )
            )

            diff_report_rows.extend(
                self._build_diff_report_rows_for_key(entity, business_key, selected_row, key_rows)
            )
        return merged_rows, by_business_key_counts, diff_report_rows

    def _stand_sort_rank_for_duplicates(self) -> dict[str, int]:
        """Приоритет стендов для выбора дублей: PROM -> PSI -> IFT."""
        preferred_order: list[str] = ["PROM", "PSI", "IFT"]
        for stand in self.stands:
            if stand not in preferred_order:
                preferred_order.append(stand)
        return {stand: idx for idx, stand in enumerate(preferred_order)}

    def _select_priority_row_for_key(self, key_rows: list[ParsedRow]) -> ParsedRow:
        """Выбирает первую строку в стенде с максимальным приоритетом."""
        rank = self._stand_sort_rank_for_duplicates()
        return min(
            key_rows,
            key=lambda row: (
                rank.get(row.stand, 999),
                row.row_num,
            ),
        )

    def _split_stands_by_selected_row(
        self, selected_row: ParsedRow, key_rows: list[ParsedRow]
    ) -> tuple[list[str], list[str]]:
        """Возвращает стенды с идентичной строкой и стенды с тем же ключом, но иными данными."""
        rank = self._stand_sort_rank_for_duplicates()
        same: set[str] = set()
        diff: set[str] = set()
        for row in key_rows:
            if row.data == selected_row.data:
                same.add(row.stand)
            else:
                diff.add(row.stand)
        return (
            sorted(same, key=lambda stand: rank.get(stand, 999)),
            sorted(diff, key=lambda stand: rank.get(stand, 999)),
        )

    def _build_diff_report_rows_for_key(
        self,
        entity: str,
        business_key: str,
        selected_row: ParsedRow,
        key_rows: list[ParsedRow],
    ) -> list[dict[str, str]]:
        """Формирует детализацию дублей по ключу для отдельного листа DIFF_REPORT."""
        diagnostics_cfg = self.config.get("excel", {}).get("diff_report_sheet", {})
        left_context = int(diagnostics_cfg.get("left_context_chars", 10))
        right_context = int(diagnostics_cfg.get("right_context_chars", 20))
        rows: list[dict[str, str]] = []
        for row in key_rows:
            diff_columns = self._find_row_diff_columns(selected_row.data, row.data)
            if not diff_columns:
                continue
            diff_positions, diff_snippets = self._build_row_diff_details(
                left=selected_row.data,
                right=row.data,
                diff_columns=diff_columns,
                left_context=left_context,
                right_context=right_context,
            )
            rows.append(
                {
                    "entity": entity,
                    "business_key": business_key,
                    "selected_stand": selected_row.stand,
                    "selected_row_num": str(selected_row.row_num),
                    "candidate_stand": row.stand,
                    "candidate_row_num": str(row.row_num),
                    "relation": "SAME_KEY_DIFFERENT_ROW",
                    "diff_columns": ",".join(diff_columns),
                    "diff_positions": diff_positions,
                    "diff_snippets": diff_snippets,
                }
            )
        return rows

    def _find_row_diff_columns(self, left: dict[str, str], right: dict[str, str]) -> list[str]:
        """Возвращает колонки, значения которых отличаются для пары строк."""
        columns = sorted(set(left.keys()) | set(right.keys()))
        diff_columns: list[str] = []
        for col in columns:
            if str(left.get(col, "")) != str(right.get(col, "")):
                diff_columns.append(col)
        return diff_columns

    def _build_row_diff_details(
        self,
        left: dict[str, str],
        right: dict[str, str],
        diff_columns: list[str],
        left_context: int,
        right_context: int,
    ) -> tuple[str, str]:
        """Формирует позиции и фрагменты отличий для пары строк."""
        if not diff_columns:
            return "", ""
        positions: list[str] = []
        snippets: list[str] = []
        for col in diff_columns:
            left_val = str(left.get(col, ""))
            right_val = str(right.get(col, ""))
            pos = self._first_diff_pos(left_val, right_val)
            positions.append(f"{col}:{pos}")
            left_part = left_val[max(0, pos - left_context) : pos + right_context]
            right_part = right_val[max(0, pos - left_context) : pos + right_context]
            snippets.append(f"{col} => [{left_part}] != [{right_part}]")
        return " | ".join(positions), " | ".join(snippets)

    def _build_cluster_payload(
        self, reference_row: ParsedRow, cluster_rows: list[ParsedRow]
    ) -> dict[str, str]:
        """
        Собирает итоговые значения кластера:
        - база из эталонной строки;
        - недостающие/пустые поля дополняются из остальных строк кластера.
        """
        payload = dict(reference_row.data)
        for row in cluster_rows:
            for field, value in row.data.items():
                if field not in payload:
                    payload[field] = value
                    continue
                if payload[field] == "" and value != "":
                    payload[field] = value
        return payload

    def _row_compatible_with_cluster(
        self,
        candidate: ParsedRow,
        cluster: list[ParsedRow],
    ) -> bool:
        for row in cluster:
            if not self._rows_compatible(left=row.data, right=candidate.data):
                return False
        return True

    def _rows_compatible(
        self,
        left: dict[str, str],
        right: dict[str, str],
    ) -> bool:
        common_fields = set(left.keys()) & set(right.keys())
        for field in common_fields:
            if left.get(field, "") != right.get(field, ""):
                return False
        return True

    def _add_diff_diagnostics(self, merged_rows: list[MergedRow]) -> None:
        """Добавляет в merged-данные колонки с детализацией отличий по одинаковому ключу."""
        by_key_rows: dict[str, list[MergedRow]] = defaultdict(list)
        service_fields = set(self.config["excel"]["formatting_defaults"].get("service_fields", []))
        service_fields.update({"diff_columns", "diff_positions", "diff_snippets"})
        for row in merged_rows:
            by_key_rows[row.business_key].append(row)

        for key, rows in by_key_rows.items():
            if len(rows) <= 1:
                for row in rows:
                    row.merged_data["diff_columns"] = ""
                    row.merged_data["diff_positions"] = ""
                    row.merged_data["diff_snippets"] = ""
                continue

            compare_columns = self._find_diff_columns(rows, service_fields)
            diff_positions, diff_snippets = self._build_diff_details(rows, compare_columns)
            for row in rows:
                row.merged_data["diff_columns"] = ",".join(compare_columns)
                row.merged_data["diff_positions"] = diff_positions
                row.merged_data["diff_snippets"] = diff_snippets

    def _find_diff_columns(
        self, rows: list[MergedRow], excluded_columns: set[str]
    ) -> list[str]:
        """Возвращает колонки, в которых есть хотя бы два разных значения."""
        columns: set[str] = set()
        for row in rows:
            columns.update(row.merged_data.keys())
        diff_columns: list[str] = []
        for col in sorted(columns):
            if col in excluded_columns:
                continue
            values = {row.merged_data.get(col, "") for row in rows}
            if len(values) > 1:
                diff_columns.append(col)
        return diff_columns

    def _build_diff_details(
        self, rows: list[MergedRow], diff_columns: list[str]
    ) -> tuple[str, str]:
        """Возвращает строку позиций отличий и короткие фрагменты."""
        if len(rows) < 2 or not diff_columns:
            return "", ""
        diagnostics_cfg = self.config.get("excel", {}).get("diff_report_sheet", {})
        left_context = int(diagnostics_cfg.get("left_context_chars", 10))
        right_context = int(diagnostics_cfg.get("right_context_chars", 20))
        left = rows[0].merged_data
        right = rows[1].merged_data
        positions: list[str] = []
        snippets: list[str] = []
        for col in diff_columns:
            left_val = str(left.get(col, ""))
            right_val = str(right.get(col, ""))
            pos = self._first_diff_pos(left_val, right_val)
            positions.append(f"{col}:{pos}")
            left_part = left_val[max(0, pos - left_context) : pos + right_context]
            right_part = right_val[max(0, pos - left_context) : pos + right_context]
            snippets.append(f"{col} => [{left_part}] != [{right_part}]")
        return " | ".join(positions), " | ".join(snippets)

    def _first_diff_pos(self, left: str, right: str) -> int:
        """Находит позицию первого отличия двух строк (0-based)."""
        min_len = min(len(left), len(right))
        for idx in range(min_len):
            if left[idx] != right[idx]:
                return idx
        return min_len

    def _export_excel(self, merged: dict[str, list[MergedRow]]) -> Path:
        """Создает Excel-файл: 1 лист на сущность + служебные колонки."""
        output_prefix = self.config["excel"]["output_name_prefix"]
        timestamp_format = self.config["excel"]["output_timestamp_format"]
        timestamp_value = datetime.now().strftime(timestamp_format)
        file_name = f"{output_prefix}_{timestamp_value}.xlsx"
        excel_path = self.paths["output_excel_dir"] / file_name
        diff_report_csv_path = (
            self.paths["output_excel_dir"] / f"{output_prefix}_{timestamp_value}_DIFF_REPORT.csv"
        )
        cons_report_csv_path = (
            self.paths["output_excel_dir"] / f"{output_prefix}_{timestamp_value}_CONS_REPORT.csv"
        )
        if self.dry_run:
            self.logger.info("Dry-run: формирование Excel пропущено, целевой путь=%s", excel_path)
            self.logger.info(
                "Dry-run: формирование CSV для DIFF_REPORT пропущено, целевой путь=%s",
                diff_report_csv_path,
            )
            self.logger.info(
                "Dry-run: формирование CSV для CONS_REPORT пропущено, целевой путь=%s",
                cons_report_csv_path,
            )
            return excel_path

        workbook = Workbook()
        workbook.remove(workbook.active)

        summary = workbook.create_sheet("SUMMARY")
        summary_headers = ["entity", "rows_total", "rows_equal_all", "rows_different"]
        summary.append(summary_headers)
        self._apply_sheet_layout(
            summary,
            self.config["excel"].get("summary_sheet", {"freeze_panes": "B2", "auto_filter_header": True}),
            len(summary_headers),
        )
        self._format_sheet(
            summary,
            base_headers=summary_headers,
            sheet_config=self.config["excel"].get("summary_sheet", {}),
            is_entity_sheet=False,
            entity_name=None,
        )

        for entity in self.entities.keys():
            sheet = workbook.create_sheet(entity)
            rows = merged.get(entity, [])
            base_headers = self._build_entity_headers(entity, rows)
            sheet.append(base_headers)

            equal_all = 0
            for item in rows:
                if item.is_equal_all:
                    equal_all += 1
                sheet.append([item.merged_data.get(header, "") for header in base_headers])

            sheet_config = self.entities[entity].get(
                "excel_sheet", {"freeze_panes": "B2", "auto_filter_header": True}
            )
            self._apply_sheet_layout(sheet, sheet_config, len(base_headers))
            self._format_sheet(
                sheet,
                base_headers=base_headers,
                sheet_config=sheet_config,
                is_entity_sheet=True,
                entity_name=entity,
            )
            summary.append([entity, len(rows), equal_all, len(rows) - equal_all])

        if self.config["excel"].get("diff_report_sheet", {}).get("enabled", True):
            self._append_diff_report_sheet(workbook, merged)

        cc = self.config.get("consistency_checks") or {}
        if is_consistency_checks_enabled(cc):
            cons_sheet = append_consistency_sheet(
                workbook,
                cc,
                self._consistency_violations,
                list(self.stands),
            )
            cons_cfg = self.config["excel"].get(
                "consistency_sheet",
                {"freeze_panes": "A2", "auto_filter_header": True},
            )
            headers = [
                str(cons_sheet.cell(row=1, column=i).value or "")
                for i in range(1, cons_sheet.max_column + 1)
            ]
            self._apply_sheet_layout(cons_sheet, cons_cfg, len(headers))
            self._format_sheet(
                cons_sheet,
                base_headers=headers,
                sheet_config=cons_cfg,
                is_entity_sheet=False,
                entity_name=None,
            )
            self._append_cons_report_sheet(workbook)

        workbook.save(excel_path)
        if self.config["excel"].get("diff_report_sheet", {}).get("export_csv", False):
            self._export_diff_report_csv(diff_report_csv_path)
        if self.config["excel"].get("cons_report_sheet", {}).get("export_csv", False):
            self._export_cons_report_csv(cons_report_csv_path)
        self.logger.info("Excel сформирован: %s", excel_path)
        return excel_path

    def _export_diff_report_csv(self, csv_path: Path) -> None:
        """Сохраняет лист DIFF_REPORT в отдельный CSV-файл для анализа."""
        headers = [
            "entity",
            "business_key",
            "selected_stand",
            "selected_row_num",
            "candidate_stand",
            "candidate_row_num",
            "relation",
            "diff_columns",
            "diff_positions",
            "diff_snippets",
        ]
        with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file, delimiter=";")
            writer.writerow(headers)
            for item in self._diff_report_rows:
                writer.writerow(
                    [
                        item.get("entity", ""),
                        item.get("business_key", ""),
                        item.get("selected_stand", ""),
                        item.get("selected_row_num", ""),
                        item.get("candidate_stand", ""),
                        item.get("candidate_row_num", ""),
                        item.get("relation", ""),
                        item.get("diff_columns", ""),
                        item.get("diff_positions", ""),
                        item.get("diff_snippets", ""),
                    ]
                )
        self.logger.info("CSV DIFF_REPORT сформирован: %s", csv_path)

    def _export_cons_report_csv(self, csv_path: Path) -> None:
        """Сохраняет лист CONS_REPORT в отдельный CSV-файл для анализа."""
        headers = [
            "entity",
            "rule_id",
            "rule_type",
            "scope",
            "stand",
            "row_num",
            "business_key",
            "severity",
            "message",
            "diff_columns",
            "diff_positions",
            "diff_snippets",
        ]
        with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file, delimiter=";")
            writer.writerow(headers)
            for violation in self._consistency_violations:
                diff_columns, diff_positions, diff_snippets = self._build_consistency_diff_details(
                    entity=str(violation.entity),
                    message=str(violation.message),
                )
                writer.writerow(
                    [
                        str(violation.entity),
                        str(violation.rule_id),
                        str(violation.rule_type),
                        str(violation.scope),
                        str(violation.stand or ""),
                        str(violation.row_num or ""),
                        str(violation.business_key or ""),
                        str(violation.severity),
                        str(violation.message),
                        diff_columns,
                        diff_positions,
                        diff_snippets,
                    ]
                )
        self.logger.info("CSV CONS_REPORT сформирован: %s", csv_path)

    def _append_diff_report_sheet(
        self, workbook: Workbook, merged: dict[str, list[MergedRow]]
    ) -> None:
        """Добавляет лист со сводкой конфликтов между стендами."""
        sheet_cfg = self.config["excel"].get(
            "diff_report_sheet",
            {"name": "DIFF_REPORT", "freeze_panes": "A2", "auto_filter_header": True},
        )
        sheet_name = sheet_cfg.get("name", "DIFF_REPORT")
        sheet = workbook.create_sheet(sheet_name)
        headers = [
            "entity",
            "business_key",
            "selected_stand",
            "selected_row_num",
            "candidate_stand",
            "candidate_row_num",
            "relation",
            "diff_columns",
            "diff_positions",
            "diff_snippets",
        ]
        sheet.append(headers)
        for item in self._diff_report_rows:
            sheet.append(
                [
                    item.get("entity", ""),
                    item.get("business_key", ""),
                    item.get("selected_stand", ""),
                    item.get("selected_row_num", ""),
                    item.get("candidate_stand", ""),
                    item.get("candidate_row_num", ""),
                    item.get("relation", ""),
                    item.get("diff_columns", ""),
                    item.get("diff_positions", ""),
                    item.get("diff_snippets", ""),
                ]
            )
        self._apply_sheet_layout(sheet, sheet_cfg, len(headers))
        self._format_sheet(
            sheet,
            base_headers=headers,
            sheet_config=sheet_cfg,
            is_entity_sheet=False,
            entity_name=None,
        )

    def _append_cons_report_sheet(self, workbook: Workbook) -> None:
        """Добавляет агрегированный лист по нарушениям консистентности."""
        sheet = workbook.create_sheet("CONS_REPORT")
        headers = [
            "entity",
            "rule_id",
            "rule_type",
            "scope",
            "stand",
            "row_num",
            "business_key",
            "severity",
            "message",
            "diff_columns",
            "diff_positions",
            "diff_snippets",
        ]
        sheet.append(headers)
        for violation in self._consistency_violations:
            diff_columns, diff_positions, diff_snippets = self._build_consistency_diff_details(
                entity=str(violation.entity),
                message=str(violation.message),
            )
            sheet.append(
                [
                    str(violation.entity),
                    str(violation.rule_id),
                    str(violation.rule_type),
                    str(violation.scope),
                    str(violation.stand or ""),
                    str(violation.row_num or ""),
                    str(violation.business_key or ""),
                    str(violation.severity),
                    str(violation.message),
                    diff_columns,
                    diff_positions,
                    diff_snippets,
                ]
            )
        sheet_cfg = {"freeze_panes": "A2", "auto_filter_header": True}
        self._apply_sheet_layout(sheet, sheet_cfg, len(headers))
        self._format_sheet(
            sheet,
            base_headers=headers,
            sheet_config=sheet_cfg,
            is_entity_sheet=False,
            entity_name=None,
        )

    def _build_consistency_diff_details(self, entity: str, message: str) -> tuple[str, str, str]:
        """Возвращает диагностику по сообщению консистентности в формате DIFF_REPORT."""
        if not message:
            return "", "", ""
        candidate_columns = self._guess_columns_from_message(entity, message)
        if not candidate_columns:
            snippet = message[:200]
            return "MESSAGE", "MESSAGE:0", f"MESSAGE => [{snippet}]"

        positions: list[str] = []
        snippets: list[str] = []
        for column in candidate_columns:
            pos = message.find(column)
            if pos < 0:
                continue
            positions.append(f"{column}:{pos}")
            left = max(0, pos - 20)
            right = min(len(message), pos + len(column) + 40)
            snippets.append(f"{column} => [{message[left:right]}]")

        if not positions:
            snippet = message[:200]
            return "MESSAGE", "MESSAGE:0", f"MESSAGE => [{snippet}]"

        return ",".join(candidate_columns), " | ".join(positions), " | ".join(snippets)

    def _guess_columns_from_message(self, entity: str, message: str) -> list[str]:
        """Пытается извлечь имена колонок из текста нарушения консистентности."""
        columns: list[str] = []
        known_headers: list[str] = []
        for stand_headers in self.entity_field_orders.get(entity, {}).values():
            known_headers.extend(stand_headers)

        unique_headers = sorted(set(known_headers), key=len, reverse=True)
        for header in unique_headers:
            if header and header in message:
                columns.append(header)
            if len(columns) >= 5:
                break
        return columns

    def _build_entity_headers(self, entity: str, rows: list[MergedRow]) -> list[str]:
        """Строит порядок колонок: эталон PROM + доп.поля в местах появления."""
        service_fields_order = self.config["excel"]["formatting_defaults"].get("service_fields", [])
        for mandatory_field in ("same_row_stands", "same_key_diff_stands"):
            if mandatory_field not in service_fields_order:
                service_fields_order.append(mandatory_field)
        present_service = [field for field in service_fields_order if any(field in row.merged_data for row in rows)]

        field_orders = self.entity_field_orders.get(entity, {})
        reference_stand = self.entities[entity].get("optional_fields", {}).get("reference_stand", "PROM")
        reference_headers = list(field_orders.get(reference_stand, []))
        combined_headers = list(reference_headers)

        for stand in self.stands:
            stand_headers = field_orders.get(stand, [])
            for index, field in enumerate(stand_headers):
                if field in combined_headers:
                    continue
                insert_index = min(index, len(combined_headers))
                combined_headers.insert(insert_index, field)

        if not combined_headers and rows:
            combined_headers = [key for key in rows[0].merged_data.keys() if key not in service_fields_order]

        for service_field in present_service:
            if service_field not in combined_headers:
                combined_headers.append(service_field)

        consistency_keys: set[str] = set()
        for row in rows:
            for k in row.merged_data.keys():
                if k.startswith("CONSIST") or k.startswith("CC_"):
                    consistency_keys.add(k)
        for ck in sorted(consistency_keys):
            if ck not in combined_headers:
                combined_headers.append(ck)
        return combined_headers

    def _apply_sheet_layout(
        self, sheet: Any, sheet_config: dict[str, Any], header_columns_count: int
    ) -> None:
        """Применяет закрепление и автофильтр для листа Excel."""
        freeze_panes = sheet_config.get("freeze_panes")
        if freeze_panes:
            sheet.freeze_panes = freeze_panes

        if sheet_config.get("auto_filter_header", False) and header_columns_count > 0:
            sheet.auto_filter.ref = f"A1:{self._column_name(header_columns_count)}1"

    def _column_name(self, column_number: int) -> str:
        """Преобразует номер колонки в Excel-имя (1 -> A, 27 -> AA)."""
        result = ""
        current = column_number
        while current > 0:
            current, remainder = divmod(current - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _format_sheet(
        self,
        sheet: Any,
        base_headers: list[str],
        sheet_config: dict[str, Any],
        is_entity_sheet: bool,
        entity_name: str | None = None,
    ) -> None:
        """Применяет форматирование листа по параметрам из config.json."""
        formatting_defaults = self.config["excel"].get("formatting_defaults", {})
        sheet_formatting = dict(formatting_defaults)
        sheet_formatting.update(sheet_config.get("formatting", {}))

        max_column_width = sheet_formatting.get("max_column_width", 150)
        header_font = Font(
            bold=sheet_formatting.get("header_bold", True),
        )
        header_alignment = Alignment(
            horizontal=sheet_formatting.get("header_horizontal", "center"),
            vertical=sheet_formatting.get("header_vertical", "center"),
            wrap_text=sheet_formatting.get("wrap_text_all_cells", True),
        )
        data_alignment = Alignment(
            horizontal=sheet_formatting.get("data_horizontal", "left"),
            vertical=sheet_formatting.get("data_vertical", "center"),
            wrap_text=sheet_formatting.get("wrap_text_all_cells", True),
        )

        yellow_fill = PatternFill(
            fill_type="solid",
            fgColor=sheet_formatting.get("missing_prom_fill", "FFFDE9"),
        )
        red_fill = PatternFill(
            fill_type="solid",
            fgColor=sheet_formatting.get("same_key_diff_fill", "FDE9E9"),
        )
        key_header_fill = PatternFill(
            fill_type="solid",
            fgColor=sheet_formatting.get("key_header_fill", "FCE4B2"),
        )
        extra_header_fill = PatternFill(
            fill_type="solid",
            fgColor=sheet_formatting.get("extra_header_fill", "E2F0D9"),
        )
        fill_priority: list[str] = sheet_formatting.get(
            "fill_priority", ["same_key_diff", "missing_prom"]
        )
        borders_cfg = sheet_formatting.get("borders", {})
        border_color = borders_cfg.get("color", "666666")
        horizontal_side = Side(
            style=borders_cfg.get("horizontal_style", "dashed"),
            color=border_color,
        )
        vertical_side = Side(
            style=borders_cfg.get("vertical_style", "dotted"),
            color=border_color,
        )
        header_separator_side = Side(
            style=borders_cfg.get("header_separator_style", "double"),
            color=border_color,
        )
        service_separator_side = Side(
            style=borders_cfg.get("service_separator_style", "thick"),
            color=border_color,
        )
        service_separator_edge = borders_cfg.get("service_separator_side", "top")
        service_fields = set(sheet_formatting.get("service_fields", []))
        for name in base_headers:
            if name.startswith("CC_") or name.startswith("CONSIST"):
                service_fields.add(name)
        service_column_indexes = {
            idx + 1 for idx, name in enumerate(base_headers) if name in service_fields
        }
        first_service_column_index = (
            min(service_column_indexes) if service_column_indexes else None
        )

        # Заголовок: жирный + центрирование.
        key_fields: set[str] = set()
        extra_fields: set[str] = set()
        if entity_name and entity_name in self.entities:
            key_fields = set(self.entities[entity_name].get("business_key", []))
            extra_fields = self.entity_extra_fields.get(entity_name, set())
        for col_idx, _ in enumerate(base_headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = Border(bottom=header_separator_side)
            if base_headers[col_idx - 1] in key_fields:
                cell.fill = key_header_fill
            elif base_headers[col_idx - 1] in extra_fields:
                cell.fill = extra_header_fill

        # Данные: выравнивание, перенос и цветовые правила.
        source_stands_index = (
            base_headers.index("source_stands") + 1 if "source_stands" in base_headers else None
        )
        same_key_diff_index = (
            base_headers.index("same_key_diff_values_flag") + 1
            if "same_key_diff_values_flag" in base_headers
            else None
        )

        for row_idx in range(2, sheet.max_row + 1):
            row_fill: PatternFill | None = self._resolve_row_fill(
                row_idx=row_idx,
                sheet=sheet,
                is_entity_sheet=is_entity_sheet,
                source_stands_index=source_stands_index,
                same_key_diff_index=same_key_diff_index,
                fill_priority=fill_priority,
                yellow_fill=yellow_fill,
                red_fill=red_fill,
            )

            for col_idx in range(1, len(base_headers) + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.alignment = data_alignment
                cell.border = Border(
                    left=vertical_side,
                    right=vertical_side,
                    top=horizontal_side,
                    bottom=horizontal_side,
                )
                if row_fill is not None:
                    cell.fill = row_fill

                if first_service_column_index is not None and col_idx == first_service_column_index:
                    cell.border = self._with_border_side(
                        border=cell.border,
                        side_name=service_separator_edge,
                        side_value=service_separator_side,
                    )

        # Автоподбор ширины колонок с ограничением max_column_width.
        for col_idx in range(1, len(base_headers) + 1):
            max_len = 0
            for row_idx in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is None:
                    continue
                value_as_text = str(cell_value)
                candidate_len = max(len(line) for line in value_as_text.splitlines()) if value_as_text else 0
                if candidate_len > max_len:
                    max_len = candidate_len
            width = min(max(max_len + 2, 10), max_column_width)
            sheet.column_dimensions[self._column_name(col_idx)].width = width

    def _with_border_side(
        self, border: Border, side_name: str, side_value: Side
    ) -> Border:
        """Возвращает Border с замененной одной стороной."""
        sides = {
            "left": border.left,
            "right": border.right,
            "top": border.top,
            "bottom": border.bottom,
        }
        if side_name not in sides:
            return border
        sides[side_name] = side_value
        return Border(
            left=sides["left"],
            right=sides["right"],
            top=sides["top"],
            bottom=sides["bottom"],
        )

    def _resolve_row_fill(
        self,
        row_idx: int,
        sheet: Any,
        is_entity_sheet: bool,
        source_stands_index: int | None,
        same_key_diff_index: int | None,
        fill_priority: list[str],
        yellow_fill: PatternFill,
        red_fill: PatternFill,
    ) -> PatternFill | None:
        """Возвращает цвет строки в порядке приоритета из config."""
        if not is_entity_sheet:
            return None

        conditions: dict[str, bool] = {}
        if same_key_diff_index is not None:
            same_key_diff_value = str(sheet.cell(row=row_idx, column=same_key_diff_index).value or "")
            conditions["same_key_diff"] = same_key_diff_value == "1"
        if source_stands_index is not None:
            stands_value = str(sheet.cell(row=row_idx, column=source_stands_index).value or "")
            has_prom = "PROM" in stands_value
            has_ift_or_psi = ("IFT" in stands_value) or ("PSI" in stands_value)
            conditions["missing_prom"] = (not has_prom) and has_ift_or_psi

        for rule_name in fill_priority:
            if conditions.get(rule_name, False):
                if rule_name == "same_key_diff":
                    return red_fill
                if rule_name == "missing_prom":
                    return yellow_fill
        return None

    def _log_found_files_summary(self, files: list[FileDescriptor]) -> None:
        """Показывает в консоли, сколько и каких файлов найдено."""
        counts_by_entity: dict[str, int] = defaultdict(int)
        for item in files:
            counts_by_entity[item.entity] += 1
        self.logger.info("Старт пайплайна run_id=%s", self.run_id)
        for entity in sorted(counts_by_entity.keys()):
            self.logger.info("Найдено файлов %s: %s", entity, counts_by_entity[entity])

    def _log_parse_summary(self, parse_stats: dict[str, Any]) -> None:
        """Пишет статистику загрузки и обновления БД."""
        self.logger.info("Файлов обработано: %s", parse_stats["total_files"])
        self.logger.info(
            "Обновление БД: новых файлов=%s, совпали хэши (пропуск)=%s",
            parse_stats["db_new_files"],
            parse_stats["db_skipped_files"],
        )
        self.logger.info("Всего строк прочитано: %s", parse_stats["processed_rows"])

    def _log_merge_summary(self, merge_stats: dict[str, Any]) -> None:
        """Пишет статистику совпадений и разночтений между стендами."""
        self.logger.info(
            "Итог merged-строк: %s (совпали во всех стендах=%s, разночтения=%s)",
            merge_stats["merged_total"],
            merge_stats["equal_all_total"],
            merge_stats["different_total"],
        )
        self.logger.info(
            "Найдено ключей с разными значениями между стендами: %s",
            merge_stats["same_key_different_values_total"],
        )
        for entity in sorted(merge_stats["by_entity"].keys()):
            item = merge_stats["by_entity"][entity]
            self.logger.info(
                "%s -> merged=%s, equal_all=%s, different=%s, same_key_diff_values=%s",
                entity,
                item["merged_rows"],
                item["equal_all_rows"],
                item["different_rows"],
                item["same_key_different_values"],
            )

    def _log_stage_timings(self, stage_timings: dict[str, float], run_started: float) -> None:
        """Логирует профилирование этапов выполнения пайплайна."""
        if not stage_timings:
            return
        self.logger.info("Профилирование этапов (секунды):")
        for stage_name in sorted(stage_timings.keys()):
            self.logger.info(" - %s: %.3f", stage_name, stage_timings[stage_name])
        self.logger.info(" - total_run: %.3f", time.perf_counter() - run_started)

